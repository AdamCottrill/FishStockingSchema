

# read in our data from the target spreadsheet:

from openpyxl import load_workbook

MAX_UPLOAD_EVENT_COUNT = 100
DATA_WORKSHEET_NAME = "DATA_TEMPLATE"

REQUIRED_FIELDS = [
"GLFSD_Stock_ID",
"Your_Agency_Stock_ID",
"AGENCY",
"LAKE",
"STATE_PROV",
"STAT_DIST",
"LS_MGMT",
"GRID_10MIN",
"LOCATION_PRIMARY",
"LOCATION_SECONDARY",
"LATITUDE",
"LONGITUDE",
"YEAR",
"MONTH",
"DAY",
"STOCK_METHOD",
"SPECIES",
"STRAIN",
"YEAR-CLASS",
"LIFE_STAGE",
"AGE_MONTHS",
"CLIP",
"CLIP_EFFICIENCY",
"PHYS-CHEM_MARK",
"TAG_TYPE",
"CWT_Number",
"TAG_RETENTION",
"MEAN_LENGTH_MM",
"TOTAL_WEIGHT_KG",
"STOCKING_MORTALITY",
"LOT_CODE",
"HATCHERY",
"NUMBER_STOCKED",
"NOTES"



]


def find_header_row(data_file, worksheet_name="DATA_TEMPLATE"):
    """Iterate over our uploaded spreadsheet - find the index of the row that contains
    the field names:

    """

    if hasattr(data_file, "open"):
        wb = load_workbook(filename=data_file.open(), data_only=True)
    else:
        wb = load_workbook(filename=data_file, data_only=True)
    ws = wb[worksheet_name]

    required_fields = {x.upper() for x in REQUIRED_FIELDS}

    for i, row in enumerate(ws.iter_rows(min_row=0, max_row=20, values_only=True)):
        vals = {x for x in row}
        shared = vals.intersection(required_fields)
        # assume that temmplate must have at least 10 fields in common
        # just incase the documentation or examples have some matching contents:
        if len(shared) > 10:
            return i + 1
    return -9



def xls2dicts(data_file):
    """A helper function to read our excel file and return a list of
    dictionaries for each row in the spreadsheet.  The keys of the
    dictionaries are determined by the first row in the spreadsheet.

    """

    # process just one more than we need - keeps us from reading too many
    # records into memory if they are going to be flagged anyway.
    # add to max count - one for the header row, and one to trip the
    # too many rows flag
    #maxrows = settings.MAX_UPLOAD_EVENT_COUNT + 2
    maxrows = MAX_UPLOAD_EVENT_COUNT + 2
    data_sheet_name = DATA_WORKSHEET_NAME

    key_row = find_header_row(data_file, data_sheet_name)

    if hasattr(data_file, "open"):
        wb = load_workbook(filename=data_file.open(), data_only=True)
    else:
        wb = load_workbook(filename=data_file, data_only=True)
    ws = wb[data_sheet_name]

    data = []

    for i, row in enumerate(
        ws.iter_rows(min_row=key_row, max_row=(maxrows + key_row), values_only=True)
    ):

        if i == 0:
            # map the values from the spread sheet to our fields
            # if it can't be found in our map - just use the value
            #keys = [xlsFields2Fdviz.get(x, x) for x in row]
            keys = row
        else:
            vals = [x for x in row]
            # if we have a blank row - stop
            if all(v is None for v in vals):
                break
            else:
                tmp = {k: v for k, v in zip(keys, vals)}
                data.append(tmp)

    return data
